from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
QA = ROOT / "outputs" / "qa"
OUT_DIR = QA / "manual_jp_zh_check"
OUT_XLSX = OUT_DIR / "MuvLuv_TDA_JP_CN_check.xlsx"


JP_RE = re.compile(r"[\u3040-\u30ff]")
MOJIBAKE_RE = re.compile(r"[□�]")
CONTROL_RE = re.compile(r"\\[A-Za-z][A-Za-z0-9_]*(?:\[[^\]]*\])?")
BAD_TERMS = [
    "地表飞行员",
    "美国地表飞行员",
    "美国飞行员",
    "飞行员们",
    "希望亡命",
    "靖国神社",
]


HEADERS = [
    "编号",
    "脚本文件",
    "ID",
    "日文原文",
    "当前中文",
    "英文参考（隐藏，不参与对齐）",
    "检查提示",
]


def clean_visible(text: str) -> str:
    text = text or ""
    text = CONTROL_RE.sub("", text)
    text = text.replace("\\n", "").replace("\\r", "").replace("\\t", "")
    return text.strip()


def excel_safe(text: str) -> str:
    text = text or ""
    return ILLEGAL_CHARACTERS_RE.sub(lambda m: f"<0x{ord(m.group(0)):02X}>", text)


def is_meaningful_jp(row: dict[str, str]) -> bool:
    jp = clean_visible(row.get("jp", ""))
    if not jp:
        return False
    if jp == "Pg9":
        return False
    return bool(JP_RE.search(jp) or re.search(r"[\u4e00-\u9fff]", jp))


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def row_key(row: dict[str, str]) -> tuple[str, str]:
    return Path(row["file"]).name, row["id"]


def flags_for(original_jp: str, current_cn: str) -> str:
    visible_current = clean_visible(current_cn)
    notes: list[str] = []
    if not visible_current:
        notes.append("中文为空")
    if "*** Text ID Not Found ***" in current_cn:
        notes.append("Text ID Not Found")
    if JP_RE.search(visible_current):
        notes.append("残留日文假名")
    if MOJIBAKE_RE.search(visible_current):
        notes.append("乱码/方块")
    for term in BAD_TERMS:
        if term in visible_current:
            notes.append(f"疑似术语: {term}")
    if visible_current in {"。", "！", "？", "……", "...", "…"}:
        notes.append("只有标点")
    return "；".join(notes)


def style_sheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Microsoft YaHei", size=10)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = header_font if cell.row == 1 else body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if cell.row == 1:
                cell.fill = header_fill
    widths = {
        "A": 8,
        "B": 30,
        "C": 15,
        "D": 62,
        "E": 62,
        "F": 62,
        "G": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.column_dimensions["F"].hidden = True


def make_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "说明"
    rows = [
        ["怎么核对"],
        ["只看每个 TDA 页里的 D/E 两列：D 是游戏原始日文，E 是现在写进游戏日文槽位的中文。"],
        ["B/C 两列是定位用的脚本文件和台词 ID；同一行就代表同一个游戏文本位置。"],
        ["F 列是英文参考，默认隐藏，不参与对齐。也就是说这张表不是按英文补空或移动台词。"],
        ["G 列为空，表示自动检查没发现乱码、残留日文、Text ID Not Found、只有标点或已知坏术语。"],
        ["如果你在游戏里看到问题，可以用 C 列 ID 或 D/E 的一句话在表里搜索。"],
    ]
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 110
    ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=14)
    for cell in ws["A"]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if cell.row > 1:
            cell.font = Font(name="Microsoft YaHei", size=11)


def build_title_sheet(wb: Workbook, title: str, label: str) -> tuple[int, int]:
    original_rows = load_csv(QA / f"{title}_original_jp_baseline_slots.csv")
    current_rows = load_csv(QA / f"{title}_current_jp_baseline_slots.csv")
    current_by_key = {row_key(r): r for r in current_rows}

    ws = wb.create_sheet(label)
    ws.append(HEADERS)

    count = 0
    flagged = 0
    for original in original_rows:
        if not is_meaningful_jp(original):
            continue
        key = row_key(original)
        current = current_by_key.get(key, {})
        original_jp = original.get("jp", "")
        current_cn = current.get("jp", "")
        note = flags_for(original_jp, current_cn)
        if note:
            flagged += 1
        count += 1
        ws.append(
            [
                count,
                key[0],
                key[1],
                excel_safe(original_jp),
                excel_safe(current_cn),
                excel_safe(original.get("en", "")),
                note,
            ]
        )

    style_sheet(ws)
    return count, flagged


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    make_readme(wb)
    summaries = []
    for title, label in [("tda01", "TDA01"), ("tda02", "TDA02"), ("tda03", "TDA03")]:
        count, flagged = build_title_sheet(wb, title, label)
        summaries.append((label, count, flagged))

    wb.save(OUT_XLSX)
    print(str(OUT_XLSX))
    for label, count, flagged in summaries:
        print(f"{label}: rows={count} flagged={flagged}")


if __name__ == "__main__":
    main()

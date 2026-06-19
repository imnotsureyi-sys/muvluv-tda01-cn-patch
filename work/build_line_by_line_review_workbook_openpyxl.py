from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
QA_DIR = ROOT / "outputs" / "qa" / "line_by_line_jp_cn_review"
INPUT_JSON = QA_DIR / "review_rows.json"
OUTPUT_XLSX = QA_DIR / "MuvLuv_TDA01-03_JP_CN_逐条核对_进行中.xlsx"

HEADERS = [
    "id",
    "jp字段",
    "cn字段",
    "是否正确",
    "如果否，打算怎么修改",
    "理由",
    "章节",
    "序号",
    "脚本文件",
    "审计提示",
]


def row_values(row: dict) -> list:
    return [
        row.get("id", ""),
        row.get("jp", ""),
        row.get("cn", ""),
        row.get("是否正确", ""),
        row.get("如果否，打算怎么修改", ""),
        row.get("理由", ""),
        row.get("title", ""),
        row.get("seq", ""),
        row.get("脚本文件", ""),
        row.get("审计提示", ""),
    ]


def apply_table_style(ws, widths: list[int]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    border_side = Side(style="thin", color="D9E2F3")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_summary_rows(rows: list[dict]) -> list[list]:
    result = [
        ["Muv-Luv TDA01-03 JP/CN 逐条核对表（进行中）", "", "", "", "", ""],
        ["规则", "", "", "", "", ""],
        ["只以 jp 字段为准核对 cn 字段；不使用英文作为参考。jp 为空时 cn 必须为空。jp 是语气/标点时 cn 也必须对应语气/标点。正常台词需语义对应、术语固定、语言通顺。", "", "", "", "", ""],
        ["状态说明", "", "", "", "", ""],
        ["“待核对”表示该行还没有完成逐条人工语义判断，不是最终结论。最终版会把每行都改成“是”或“否”。", "", "", "", "", ""],
        ["", "", "", "", "", ""],
        ["章节", "总行数", "已判定是", "已判定否", "待核对", "有审计提示"],
    ]
    for title in ["TDA01", "TDA02", "TDA03"]:
        subset = [row for row in rows if row.get("title") == title]
        result.append([
            title,
            len(subset),
            sum(row.get("是否正确") == "是" for row in subset),
            sum(row.get("是否正确") == "否" for row in subset),
            sum(row.get("是否正确") == "待核对" for row in subset),
            sum(bool(row.get("审计提示")) for row in subset),
        ])
    return result


def main() -> None:
    rows = json.loads(INPUT_JSON.read_text(encoding="utf-8"))
    wb = Workbook()
    summary = wb.active
    summary.title = "说明与进度"
    for row in build_summary_rows(rows):
        summary.append(row)
    summary.merge_cells("A1:F1")
    apply_table_style(summary, [24, 12, 12, 12, 12, 14])
    summary["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary["A1"].font = Font(bold=True, color="FFFFFF")
    for cell in summary[7]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)

    for title in ["TDA01", "TDA02", "TDA03"]:
        ws = wb.create_sheet(title)
        ws.append(HEADERS)
        for row in rows:
            if row.get("title") == title:
                ws.append(row_values(row))
        apply_table_style(ws, [18, 72, 72, 12, 44, 48, 10, 9, 44, 48])

    QA_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
